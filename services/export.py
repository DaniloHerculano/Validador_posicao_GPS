"""
export.py — Geração de planilha Excel multi-aba COM gráficos nativos e CSV.
"""
import io
import re
import pandas as pd
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

from services.ui import fmt_duracao
from services.analise import resumo_raio, analisar_buffer_lifo

# Cores Stoneridge
C_RED = "DD0933"
C_SLATE = "2C3946"
C_LIGHT = "D5E2F2"


def _nome_aba(nome: str) -> str:
    limpo = re.sub(r'[\[\]\:\*\?\/\\]', '_', nome)
    return (limpo[:28] or "comp")[:31]


def nome_arquivo_seguro(nome: str) -> str:
    return re.sub(r'[^A-Za-z0-9_.-]', '_', nome)


def _estiliza_header(ws, ncols, nrow=1):
    fill = PatternFill("solid", fgColor=C_SLATE)
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    thin = Side(style="thin", color="DCE4EE")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=nrow, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)


def _escreve_df(ws, df, start_row=1):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _estiliza_header(ws, len(df.columns), start_row)
    # largura automática simples
    for i, col in enumerate(df.columns, 1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].head(50)])
        ws.column_dimensions[ws.cell(row=1, column=i).coordinate[:1]].width = min(maxlen + 3, 45)


def _comenta_colunas(ws, df, textos: dict, nrow=1):
    """
    Adiciona um comentário (balão ao passar o mouse) nos cabeçalhos das
    colunas listadas em `textos` — equivalente, no Excel, ao ícone "?" da
    plataforma. `textos` é {nome_da_coluna: explicação}.
    """
    for i, col in enumerate(df.columns, 1):
        if col in textos:
            c = ws.cell(row=nrow, column=i)
            c.comment = Comment(textos[col], "Validador de Posicionamento", width=260, height=110)


def _dtec(item):
    """Retorna o dataframe técnico completo (CSV) do equipamento, se houver."""
    return item.get("df_tecnico", item.get("df")) if item else None


GLOSSARIO = [
    ("GPS Real", "Posição obtida diretamente do receptor GPS do equipamento (gps=true)."),
    ("Posição Estimada", "Posição calculada pelo sistema (trilateração/rede) quando não há "
                          "fixação de GPS, com um raio de incerteza informado pelo próprio sistema."),
    ("Misto (Real + Estimada)", "Equipamento com parte dos registros em GPS Real e parte em "
                                 "Posição Estimada ao longo do teste."),
    ("Erro de Posição (km)", "Distância geodésica entre o ponto de referência (GPS real, "
                              "confiável) e o ponto do equipamento comparado, no mesmo instante."),
    ("Raio do Sistema (km)", "Raio de incerteza que o próprio sistema PST informa para uma "
                              "posição estimada (vem do arquivo KML). 'Dentro do raio' significa "
                              "que a posição real caiu dentro do círculo de incerteza prometido."),
    ("Tolerância de horário (min)", "Janela máxima de tempo usada para parear um registro do "
                                     "equipamento comparado com o ponto mais próximo da referência."),
    ("Tolerância de fusão CSV↔XLS (s)", "Janela de tempo usada para casar o log técnico (CSV) "
                                         "com o ponto de posição (XLS) do mesmo equipamento."),
    ("datetime_module", "Data/hora em que o registro foi gerado dentro do próprio módulo "
                         "(rastreador), antes de ser transmitido."),
    ("datetime_server", "Data/hora em que o registro chegou ao servidor."),
    ("Latência", "Tempo entre datetime_module e datetime_server — quanto tempo o registro "
                 "levou para chegar ao servidor depois de gerado no módulo."),
    ("Latência em Tempo Real", "Latência considerando apenas registros NÃO bufferizados "
                                "(bufferstatus=false). É a métrica que reflete a velocidade "
                                "real de transmissão do equipamento."),
    ("Buffer / bufferstatus=true", "Registro que o módulo guardou localmente por ter ficado "
                                    "sem sinal, e enviou depois, todo atrasado, quando a conexão "
                                    "voltou. Não indica falha de envio em tempo real — é o efeito "
                                    "esperado de uma perda de sinal."),
    ("LIFO (Last In, First Out)",
     "Comportamento esperado do módulo ao recuperar o buffer: ele sobe primeiro os "
     "registros mais recentes e só depois os mais antigos (como uma pilha, e não uma "
     "fila). É o funcionamento normal do equipamento, não um erro.\n\n"
     "Exemplo: o módulo perde sinal e guarda os pontos das 12:10, 12:20 e 12:30 "
     "(datetime_module). Ao reconectar, ele deve enviar primeiro o das 12:30, depois "
     "o das 12:20, depois o das 12:10 — os três chegam ao servidor quase juntos "
     "(datetime_server próximos), mas nessa ordem decrescente."),
    ("Episódio de Buffer",
     "Uma rajada contínua de registros bufferizados chegando ao servidor — "
     "corresponde a um ciclo de perda + recuperação de sinal. No exemplo acima, os "
     "3 pontos (12:10/12:20/12:30) que sobem juntos formam 1 episódio, com "
     "'Registros' = 3."),
    ("% Episódios em ordem LIFO",
     "Percentual de episódios em que TODOS os registros chegaram em ordem "
     "decrescente de datetime_module (mais novo primeiro) — ou seja, LIFO perfeito, "
     "sem nenhum par fora de ordem."),
    ("Correlação de ordem (rho)", "Mede a relação entre a ordem de chegada ao servidor e o "
                                   "datetime_module dentro de um episódio de buffer. "
                                   "-1 = ordem LIFO perfeita · 0 = ordem aleatória · "
                                   "+1 = ordem invertida (FIFO)."),
    ("Registros fora de ordem", "Registros que pertencem a episódios de buffer onde a ordem "
                                 "LIFO esperada não foi 100% respeitada."),
    ("% pares fora de ordem",
     "Dentro de um episódio, comparamos CADA PAR de registros pela ordem de chegada "
     "ao servidor: se o que chegou depois tem um datetime_module mais novo que o que "
     "chegou antes, esse par está 'fora de ordem'.\n\n"
     "Exemplo com 4 registros chegando nesta ordem: 08:44, 08:34, 08:22, 15:15. Os "
     "3 primeiros vêm certinho (decrescendo). Mas o 4º (15:15) é mais novo que os 3 "
     "anteriores — ele deveria ter chegado primeiro. Isso gera 3 pares fora de ordem "
     "em 6 pares possíveis = 50%."),
    ("Quebrou a ordem?",
     "Marca o registro específico que chegou com um datetime_module mais novo do que "
     "algum registro anterior no mesmo episódio — ou seja, o ponto onde a sequência "
     "LIFO esperada foi rompida. Ver a aba Buffer_LIFO_Detalhe para o registro a "
     "registro de cada episódio fora de ordem."),
    ("% ≤90s (tempo real)", "Percentual dos registros em tempo real com latência de até 90 "
                             "segundos — limite de referência considerado saudável."),
]


def _aba_glossario(wb):
    ws = wb.create_sheet("Glossário")
    ws["A1"] = "Glossário — Termos usados neste relatório"
    ws["A1"].font = Font(bold=True, size=14, color=C_SLATE)
    ws["A2"] = ("Este relatório foi pensado para ser lido de forma independente, mesmo sem "
                "acesso à plataforma. Use esta aba como referência para os termos técnicos.")
    ws["A2"].font = Font(size=10, italic=True, color="6B7F8F")
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 28
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    hdr = 4
    for j, h in enumerate(["Termo", "Explicação"], 1):
        c = ws.cell(row=hdr, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=C_SLATE)
        c.font = Font(color="FFFFFF", bold=True)
    for i, (termo, exp) in enumerate(GLOSSARIO, 1):
        r = hdr + i
        ct = ws.cell(row=r, column=1, value=termo)
        ct.font = Font(bold=True, color=C_SLATE)
        ct.alignment = Alignment(vertical="top", wrap_text=True)
        ce = ws.cell(row=r, column=2, value=exp)
        ce.alignment = Alignment(vertical="top", wrap_text=True)
        n_linhas = exp.count("\n") + 1 + len(exp) // 90
        ws.row_dimensions[r].height = max(32, 15 * n_linhas)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    return ws



def gerar_excel(df_resumo, resultados: dict, dados: list,
                raios=(1.0, 3.0, 5.0), tolerancia=None, tol_fusao=None) -> bytes:
    """
    Excel autossuficiente (pode ser lido/analisado sem acesso à plataforma):
      - Informações (capa + parâmetros do teste + período)
      - Resumo (tabela + gráfico de barras de erro médio + % por raio)
      - Rede_Bateria (consolidado + gráfico de barras de tecnologia)
      - Latência_Buffer (latência em tempo real x buffer, por equipamento)
      - Buffer_LIFO (verificação de ordem LIFO na recuperação do buffer)
      - Raio_Sistema (validação contra o raio de incerteza do KML)
      - Uma aba por equipamento sincronizado (tabela + linha de erro no tempo)
      - Glossário (explicação de todos os termos técnicos usados)
    """
    from openpyxl import Workbook
    import datetime as _dt
    wb = Workbook()
    raio1, raio2, raio3 = raios

    # ── Aba CAPA / INFORMAÇÕES ──
    wscapa = wb.active
    wscapa.title = "Informações"
    wscapa["A1"] = "TESTE DE RODAGEM — Relatório de Análise"
    wscapa["A1"].font = Font(bold=True, size=16, color=C_SLATE)
    wscapa["A2"] = "Análise comparativa GPS Real × Posição Estimada · Stoneridge Brasil"
    wscapa["A2"].font = Font(size=11, color=C_RED, bold=True)
    wscapa["A4"] = "Gerado em:"
    wscapa["B4"] = _dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    wscapa["A4"].font = Font(bold=True)

    # Período do teste (min/max datetime_module entre todos os equipamentos)
    momentos = []
    for d in dados:
        df_d = d.get("df")
        if df_d is not None and "datetime_module" in df_d.columns:
            s = df_d["datetime_module"].dropna()
            if len(s):
                momentos.append((s.min(), s.max()))
    wscapa["A5"] = "Período do teste:"
    wscapa["A5"].font = Font(bold=True)
    if momentos:
        ini = min(m[0] for m in momentos); fim = max(m[1] for m in momentos)
        wscapa["B5"] = f"{ini.strftime('%d/%m/%Y %H:%M')} — {fim.strftime('%d/%m/%Y %H:%M')}"

    # Parâmetros usados na análise (coluna H, longe da tabela de equipamentos)
    wscapa["H4"] = "Parâmetros do teste"
    wscapa["H4"].font = Font(bold=True, size=12, color=C_SLATE)
    linhas_param = []
    if tolerancia is not None:
        linhas_param.append(("Tolerância de horário", f"{tolerancia} min"))
    if tol_fusao is not None:
        linhas_param.append(("Tolerância de fusão CSV↔XLS", f"{tol_fusao} s"))
    linhas_param.append(("Raio 1 / 2 / 3", f"{raio1} / {raio2} / {raio3} km"))
    for i, (k, v) in enumerate(linhas_param):
        wscapa.cell(row=5 + i, column=8, value=k).font = Font(bold=True)
        wscapa.cell(row=5 + i, column=9, value=v)
    wscapa.column_dimensions["H"].width = 28
    wscapa.column_dimensions["I"].width = 22

    wscapa["A6"] = "Equipamentos analisados"
    wscapa["A6"].font = Font(bold=True, size=12, color=C_SLATE)
    # Tabela de equipamentos
    hdr_row = 7
    for j, h in enumerate(["Modelo", "PIN", "Arquivo", "Tipo", "Fonte", "Registros"], 1):
        c = wscapa.cell(row=hdr_row, column=j, value=h)
        c.fill = PatternFill("solid", fgColor=C_SLATE)
        c.font = Font(color="FFFFFF", bold=True)
    for i, d in enumerate(dados, 1):
        r = hdr_row + i
        vals = [d.get("modelo", "—"), d.get("pin", "—"), d.get("arquivo", "—"),
                d.get("tipo", "—"), d.get("fonte", "—"), d.get("registros", 0)]
        for j, v in enumerate(vals, 1):
            wscapa.cell(row=r, column=j, value=v)
    for col, w in zip("ABCDEF", [12, 14, 40, 20, 14, 10]):
        wscapa.column_dimensions[col].width = w

    # ── Aba RESUMO ──
    ws = wb.create_sheet("Resumo")
    if df_resumo is not None and len(df_resumo) > 0:
        _escreve_df(ws, df_resumo)
        n = len(df_resumo)

        # Gráfico de barras: erro médio (procura coluna que contenha "Erro Médio")
        col_erro = next((i for i, c in enumerate(df_resumo.columns, 1)
                         if "Erro Médio" in str(c)), None)
        col_equip = next((i for i, c in enumerate(df_resumo.columns, 1)
                          if "Equipamento" in str(c)), 1)
        if col_erro:
            bar = BarChart()
            bar.title = "Erro Médio por Equipamento (km)"
            bar.type = "col"
            bar.style = 10
            data = Reference(ws, min_col=col_erro, min_row=1, max_row=n + 1)
            cats = Reference(ws, min_col=col_equip, min_row=2, max_row=n + 1)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 8; bar.width = 16
            ws.add_chart(bar, f"A{n + 4}")

        # Gráfico de barras agrupadas: % por raio
        cols_pct = [i for i, c in enumerate(df_resumo.columns, 1) if "%" in str(c)]
        if cols_pct:
            bar2 = BarChart()
            bar2.title = "% Dentro de Cada Raio"
            bar2.type = "col"; bar2.grouping = "clustered"; bar2.style = 12
            data = Reference(ws, min_col=min(cols_pct), max_col=max(cols_pct),
                             min_row=1, max_row=n + 1)
            cats = Reference(ws, min_col=col_equip, min_row=2, max_row=n + 1)
            bar2.add_data(data, titles_from_data=True)
            bar2.set_categories(cats)
            bar2.height = 8; bar2.width = 16
            ws.add_chart(bar2, f"A{n + 22}")

    # ── Aba REDE & BATERIA ──
    linhas = []
    for d in dados:
        df = d["df"]
        total = len(df)
        bat = df["_bateria_pct"].dropna() if "_bateria_pct" in df.columns else pd.Series(dtype=float)
        tem_tech = "_tech" in df.columns
        op = (df["_operadora"].mode().iloc[0]
              if "_operadora" in df.columns and df["_operadora"].notna().any() else "—")
        linhas.append({
            "Modelo": d.get("modelo", "—"), "PIN": d["pin"], "Tipo": d["tipo"],
            "Fonte": d.get("fonte", "—"), "Registros": total,
            "4G": int(df["_tech"].str.startswith("4G").sum()) if tem_tech else 0,
            "3G": int(df["_tech"].str.startswith("3G").sum()) if tem_tech else 0,
            "2G": int(df["_tech"].str.startswith("2G").sum()) if tem_tech else 0,
            "Operadora": op,
            "Bateria Início (%)": bat.iloc[0] if len(bat) else None,
            "Bateria Fim (%)": bat.iloc[-1] if len(bat) else None,
        })
    if linhas:
        ws2 = wb.create_sheet("Rede_Bateria")
        df_rb = pd.DataFrame(linhas)
        _escreve_df(ws2, df_rb)
        n = len(df_rb)
        # Barras empilhadas 2G/3G/4G por equipamento
        c4 = list(df_rb.columns).index("4G") + 1
        c2 = list(df_rb.columns).index("2G") + 1
        cmod = list(df_rb.columns).index("Modelo") + 1
        bar = BarChart(); bar.title = "Distribuição de Tecnologia por Equipamento"
        bar.type = "col"; bar.grouping = "stacked"; bar.overlap = 100; bar.style = 10
        data = Reference(ws2, min_col=min(c4, c2), max_col=max(c4, c2), min_row=1, max_row=n + 1)
        cats = Reference(ws2, min_col=cmod, min_row=2, max_row=n + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 9; bar.width = 18
        ws2.add_chart(bar, f"A{n + 4}")

    # ── Aba LATÊNCIA & BUFFER ──
    linhas_lat = []
    for d in dados:
        dft = _dtec(d)
        if dft is None or "_latencia_s" not in dft.columns or dft["_latencia_s"].isna().all():
            continue
        dfl = dft.dropna(subset=["_latencia_s"]).copy()
        if "_buffer_bool" in dfl.columns:
            bufm = dfl["_buffer_bool"].fillna(False)
        elif "bufferstatus" in dfl.columns:
            bufm = dfl["bufferstatus"].astype(str).str.lower().isin(["t", "true", "1"])
        else:
            bufm = pd.Series(False, index=dfl.index)
        dfl_real = dfl[~bufm]; dfl_buf = dfl[bufm]
        base_real = dfl_real if len(dfl_real) > 0 else dfl
        media = base_real["_latencia_s"].mean(); mx = base_real["_latencia_s"].max()
        pok = (base_real["_latencia_s"] <= 90).mean() * 100
        linhas_lat.append({
            "Equipamento": d.get("arquivo", "—"), "Modelo": d.get("modelo", "—"),
            "Registros c/ latência": len(dfl),
            "Latência média tempo real (s)": round(media, 1),
            "Latência média tempo real": fmt_duracao(media, decimais=True),
            "Latência máx tempo real (s)": round(mx, 1),
            "Latência máx tempo real": fmt_duracao(mx),
            "% ≤90s (tempo real)": round(pok, 1),
            "Registros bufferizados": len(dfl_buf),
            "% bufferizados": round(len(dfl_buf) / len(dfl) * 100, 1) if len(dfl) else 0,
            "Buffer latência mín (s)": round(dfl_buf["_latencia_s"].min(), 1) if len(dfl_buf) else None,
            "Buffer latência máx (s)": round(dfl_buf["_latencia_s"].max(), 1) if len(dfl_buf) else None,
            "Buffer latência média (s)": round(dfl_buf["_latencia_s"].mean(), 1) if len(dfl_buf) else None,
            "Buffer latência (legível)": (
                f"{fmt_duracao(dfl_buf['_latencia_s'].min())} a "
                f"{fmt_duracao(dfl_buf['_latencia_s'].max())} "
                f"(média {fmt_duracao(dfl_buf['_latencia_s'].mean())})"
                if len(dfl_buf) else "—"),
        })
    if linhas_lat:
        ws4 = wb.create_sheet("Latência_Buffer")
        df_lat = pd.DataFrame(linhas_lat)
        _escreve_df(ws4, df_lat)
        _comenta_colunas(ws4, df_lat, {
            "Latência média tempo real (s)": "Tempo médio entre datetime_module e "
                "datetime_server, considerando só registros NÃO bufferizados (envio "
                "em tempo real).",
            "Registros bufferizados": "Registros com bufferstatus=true — guardados "
                "localmente por perda de sinal e enviados depois, em atraso.",
            "% bufferizados": "Percentual de registros que vieram do buffer em vez de "
                "tempo real. Ver aba Buffer_LIFO para a verificação de ordem.",
        })
        n = len(df_lat)
        cmed = list(df_lat.columns).index("Latência média tempo real (s)") + 1
        cpb = list(df_lat.columns).index("% bufferizados") + 1
        cequip = list(df_lat.columns).index("Equipamento") + 1
        bar_lat = BarChart(); bar_lat.title = "Latência Média em Tempo Real (s) por Equipamento"
        bar_lat.type = "col"; bar_lat.style = 10
        data = Reference(ws4, min_col=cmed, min_row=1, max_row=n + 1)
        cats = Reference(ws4, min_col=cequip, min_row=2, max_row=n + 1)
        bar_lat.add_data(data, titles_from_data=True); bar_lat.set_categories(cats)
        bar_lat.height = 8; bar_lat.width = 16
        ws4.add_chart(bar_lat, f"A{n + 4}")

        bar_buf = BarChart(); bar_buf.title = "% de Registros Bufferizados por Equipamento"
        bar_buf.type = "col"; bar_buf.style = 11
        data = Reference(ws4, min_col=cpb, min_row=1, max_row=n + 1)
        bar_buf.add_data(data, titles_from_data=True); bar_buf.set_categories(cats)
        bar_buf.height = 8; bar_buf.width = 16
        ws4.add_chart(bar_buf, f"A{n + 22}")

    # ── Aba BUFFER & LIFO ──
    resumo_lifo_linhas = []
    episodios_linhas = []
    registros_lifo_linhas = []
    for d in dados:
        dft = _dtec(d)
        if dft is None:
            continue
        lifo = analisar_buffer_lifo(dft)
        if not lifo:
            continue
        nome_eq = d.get("arquivo", "—")
        resumo_lifo_linhas.append({
            "Equipamento": nome_eq, "Episódios de buffer": lifo["n_episodios"],
            "Episódios em ordem LIFO": lifo["n_conforme"],
            "% em ordem LIFO": lifo["pct_conforme"],
            "Registros fora de ordem": lifo["registros_fora_ordem"],
            "Correlação média de ordem": lifo["rho_medio"],
        })
        ep_df = lifo["episodios_df"].copy()
        ep_df.insert(0, "Equipamento", nome_eq)
        episodios_linhas.append(ep_df)

        # Detalhe registro a registro, só dos episódios que ficaram fora de ordem
        eps_problematicos = set(ep_df.loc[~ep_df["conforme_lifo"], "episodio"])
        if eps_problematicos:
            reg_df = lifo["registros_df"]
            reg_prob = reg_df[reg_df["episodio"].isin(eps_problematicos)].copy()
            reg_prob.insert(0, "Equipamento", nome_eq)
            registros_lifo_linhas.append(reg_prob)

    if resumo_lifo_linhas:
        ws5 = wb.create_sheet("Buffer_LIFO")
        ws5["A1"] = "Verificação de Ordem LIFO na Recuperação do Buffer"
        ws5["A1"].font = Font(bold=True, size=13, color=C_SLATE)
        ws5["A2"] = ("LIFO: ao perder sinal, o módulo deve subir os registros mais recentes "
                     "primeiro ao reconectar. Ver aba Glossário para mais detalhes.")
        ws5["A2"].font = Font(size=9, italic=True, color="6B7F8F")
        df_resumo_lifo = pd.DataFrame(resumo_lifo_linhas)
        _escreve_df(ws5, df_resumo_lifo, start_row=3)
        _comenta_colunas(ws5, df_resumo_lifo, {
            "% em ordem LIFO": "Percentual de episódios de buffer em que os registros "
                "chegaram 100% em ordem decrescente de datetime_module (LIFO perfeito).",
            "Correlação média de ordem": "-1 = LIFO perfeito · 0 = ordem aleatória · "
                "+1 = ordem invertida (FIFO).",
        }, nrow=3)
        n2 = len(df_resumo_lifo)
        cequip2 = list(df_resumo_lifo.columns).index("Equipamento") + 1
        cpct2 = list(df_resumo_lifo.columns).index("% em ordem LIFO") + 1
        bar_lifo = BarChart(); bar_lifo.title = "% de Episódios em Ordem LIFO por Equipamento"
        bar_lifo.type = "col"; bar_lifo.style = 10
        data = Reference(ws5, min_col=cpct2, min_row=3, max_row=3 + n2)
        cats = Reference(ws5, min_col=cequip2, min_row=4, max_row=3 + n2)
        bar_lifo.add_data(data, titles_from_data=True); bar_lifo.set_categories(cats)
        bar_lifo.height = 8; bar_lifo.width = 16
        anchor_lifo = chr(65 + len(df_resumo_lifo.columns) + 2)
        ws5.add_chart(bar_lifo, f"{anchor_lifo}3")

        if episodios_linhas:
            cur = ws5.max_row
            ws5.cell(row=cur + 2, column=1,
                     value="Detalhe por episódio de buffer").font = Font(
                bold=True, size=12, color=C_SLATE)
            df_ep_all = pd.concat(episodios_linhas, ignore_index=True)
            df_ep_all = df_ep_all.rename(columns={
                "episodio": "Episódio", "n_registros": "Registros",
                "inicio_server": "Início (servidor)", "fim_server": "Fim (servidor)",
                "modulo_mais_novo": "Módulo mais novo", "modulo_mais_antigo": "Módulo mais antigo",
                "violacoes": "Violações", "pares_totais": "Pares totais",
                "taxa_desordem_pct": "% fora de ordem", "rho_ordem": "Correlação de ordem",
                "conforme_lifo": "Conforme LIFO?",
            })
            df_ep_all["Conforme LIFO?"] = df_ep_all["Conforme LIFO?"].map({True: "Sim", False: "Não"})
            for c in ["Início (servidor)", "Fim (servidor)", "Módulo mais novo", "Módulo mais antigo"]:
                if c in df_ep_all.columns:
                    df_ep_all[c] = df_ep_all[c].astype(str)
            det_hdr_row = cur + 3
            _escreve_df(ws5, df_ep_all, start_row=det_hdr_row)
            _comenta_colunas(ws5, df_ep_all, {
                "Episódio": "Número da rajada de recuperação (perda de sinal seguida "
                    "de reconexão). Ex.: os pontos das 12:10/12:20/12:30 que sobem "
                    "juntos ao reconectar formam 1 episódio.",
                "Registros": "Quantos pontos bufferizados vieram nessa rajada.",
                "Início (servidor)": "Quando o primeiro registro dessa rajada chegou "
                    "ao servidor.",
                "Fim (servidor)": "Quando o último registro dessa rajada chegou ao "
                    "servidor.",
                "% fora de ordem": "De todos os pares de registros da rajada, quantos "
                    "vieram na ordem errada (mais novo chegando depois do mais "
                    "antigo). Ex.: 4 registros chegam na ordem 08:44 → 08:34 → 08:22 "
                    "→ 15:15. O último (15:15) é mais novo que os 3 anteriores, "
                    "gerando 3 pares fora de ordem em 6 pares possíveis = 50%.",
                "Conforme LIFO?": "'Não' significa que pelo menos 1 par de registros "
                    "nessa rajada veio fora da ordem decrescente esperada. Ver aba "
                    "Buffer_LIFO_Detalhe para o registro que quebrou a ordem.",
            }, nrow=det_hdr_row)

    # ── Aba BUFFER & LIFO — DETALHE REGISTRO A REGISTRO ──
    if registros_lifo_linhas:
        ws5b = wb.create_sheet("Buffer_LIFO_Detalhe")
        ws5b["A1"] = "Buffer fora de ordem — registro a registro"
        ws5b["A1"].font = Font(bold=True, size=13, color=C_SLATE)
        ws5b["A2"] = (
            "Mostra, registro por registro, a ordem real de chegada ao servidor "
            "dentro de cada episódio marcado como fora de ordem LIFO. 'Quebrou a "
            "ordem?' = Sim no registro que chegou com um datetime_module mais novo "
            "do que algum registro anterior do mesmo episódio — ou seja, o ponto "
            "exato onde a sequência esperada (mais novo primeiro) foi rompida.")
        ws5b["A2"].font = Font(size=9, italic=True, color="6B7F8F")
        ws5b.merge_cells("A2:G2")
        ws5b.row_dimensions[2].height = 42
        ws5b["A2"].alignment = Alignment(wrap_text=True, vertical="top")

        df_reg_all = pd.concat(registros_lifo_linhas, ignore_index=True)
        df_reg_all = df_reg_all.rename(columns={
            "episodio": "Episódio", "ordem_chegada": "Ordem de chegada",
            "datetime_module": "Datetime módulo", "datetime_server": "Datetime servidor",
            "quebrou_ordem": "Quebrou a ordem?",
        })
        df_reg_all["Quebrou a ordem?"] = df_reg_all["Quebrou a ordem?"].map(
            {True: "Sim", False: "Não"})
        for c in ["Datetime módulo", "Datetime servidor"]:
            df_reg_all[c] = df_reg_all[c].astype(str)
        reg_hdr_row = 4
        _escreve_df(ws5b, df_reg_all, start_row=reg_hdr_row)
        _comenta_colunas(ws5b, df_reg_all, {
            "Ordem de chegada": "1º, 2º, 3º... registro a chegar ao servidor dentro "
                "desse episódio (ordenado por datetime_server).",
            "Quebrou a ordem?": "'Sim' = este registro tem datetime_module mais novo "
                "do que algum registro anterior na mesma rajada — deveria ter "
                "chegado antes, não depois.",
        }, nrow=reg_hdr_row)
        # Destaca em vermelho claro os registros que quebraram a ordem
        col_quebrou = list(df_reg_all.columns).index("Quebrou a ordem?") + 1
        fill_alerta = PatternFill("solid", fgColor="FBE1E6")
        for r in range(reg_hdr_row + 1, reg_hdr_row + 1 + len(df_reg_all)):
            if ws5b.cell(row=r, column=col_quebrou).value == "Sim":
                for c in range(1, len(df_reg_all.columns) + 1):
                    ws5b.cell(row=r, column=c).fill = fill_alerta

    # ── Aba RAIO DO SISTEMA (KML) ──
    resumo_raio_linhas = []
    fora_linhas = []
    for nome, df in resultados.items():
        if len(df) == 0 or "dentro_raio" not in df.columns:
            continue
        rr = resumo_raio(df)
        if rr is None:
            continue
        resumo_raio_linhas.append({
            "Equipamento": nome, "Pontos avaliados": rr["pontos"],
            "Dentro do raio": rr["dentro"], "% dentro do raio": rr["pct_dentro"],
            "Raio médio (km)": rr["raio_medio"], "Raio mín (km)": rr["raio_min"],
            "Raio máx (km)": rr["raio_max"],
        })
        val = df.dropna(subset=["dentro_raio", "distancia_km", "raio_km"]).copy()
        val["dentro_raio"] = val["dentro_raio"].astype(bool)
        fora_df = val[~val["dentro_raio"]].copy()
        if len(fora_df) > 0:
            fora_df["excedente_km"] = (fora_df["distancia_km"] - fora_df["raio_km"]).round(3)
            cols = [c for c in ["horario_comp", "distancia_km", "raio_km", "excedente_km", "endereco_comp"]
                    if c in fora_df.columns]
            tab = fora_df.nlargest(min(15, len(fora_df)), "excedente_km")[cols].copy()
            tab.insert(0, "Equipamento", nome)
            fora_linhas.append(tab)

    if resumo_raio_linhas:
        ws6 = wb.create_sheet("Raio_Sistema")
        ws6["A1"] = "Validação contra o Raio de Incerteza do Sistema (KML)"
        ws6["A1"].font = Font(bold=True, size=13, color=C_SLATE)
        ws6["A2"] = ("Compara a distância medida com o raio de incerteza que o próprio "
                     "sistema PST informou para cada posição estimada.")
        ws6["A2"].font = Font(size=9, italic=True, color="6B7F8F")
        df_rr = pd.DataFrame(resumo_raio_linhas)
        _escreve_df(ws6, df_rr, start_row=3)
        _comenta_colunas(ws6, df_rr, {
            "% dentro do raio": "Percentual de posições cuja distância medida (referência "
                "× amostra) ficou dentro do raio de incerteza informado pelo sistema (KML).",
        }, nrow=3)
        n3 = len(df_rr)
        cequip3 = list(df_rr.columns).index("Equipamento") + 1
        cpct3 = list(df_rr.columns).index("% dentro do raio") + 1
        bar_raio = BarChart(); bar_raio.title = "% Dentro do Raio do Sistema por Equipamento"
        bar_raio.type = "col"; bar_raio.style = 10
        data = Reference(ws6, min_col=cpct3, min_row=3, max_row=3 + n3)
        cats = Reference(ws6, min_col=cequip3, min_row=4, max_row=3 + n3)
        bar_raio.add_data(data, titles_from_data=True); bar_raio.set_categories(cats)
        bar_raio.height = 8; bar_raio.width = 16
        anchor_raio = chr(65 + len(df_rr.columns) + 2)
        ws6.add_chart(bar_raio, f"{anchor_raio}3")

        if fora_linhas:
            cur3 = ws6.max_row
            ws6.cell(row=cur3 + 2, column=1,
                     value="Piores excedentes (fora do raio do sistema)").font = Font(
                bold=True, size=12, color=C_SLATE)
            df_fora = pd.concat(fora_linhas, ignore_index=True)
            ren2 = {"horario_comp": "Horário", "distancia_km": "Distância medida (km)",
                    "raio_km": "Raio sistema (km)", "excedente_km": "Excedente (km)",
                    "endereco_comp": "Endereço estimado"}
            df_fora = df_fora.rename(columns=ren2)
            if "Horário" in df_fora.columns:
                df_fora["Horário"] = df_fora["Horário"].astype(str)
            _escreve_df(ws6, df_fora, start_row=cur3 + 3)

    # ── Abas por equipamento sincronizado ──
    for nome, df in resultados.items():
        if len(df) == 0:
            continue
        ws3 = wb.create_sheet(f"Sync_{_nome_aba(nome)}"[:31])
        cols = [c for c in ["horario_comp", "distancia_km", "raio_km", "dentro_raio",
                            "estimada_comp", "gps_valido_comp", "lat_comp", "lon_comp",
                            "_tech_comp", "_operadora_comp", "_banda_comp",
                            "endereco_comp"] if c in df.columns]
        dfx = df[cols].copy()
        # Traduzir flags booleanas para texto legível
        if "estimada_comp" in dfx.columns:
            dfx["estimada_comp"] = dfx["estimada_comp"].map(
                {True: "Estimada", False: "GPS Real"}).fillna("—")
        if "gps_valido_comp" in dfx.columns:
            dfx["gps_valido_comp"] = dfx["gps_valido_comp"].map(
                {True: "Válido", False: "Inválido"}).fillna("—")
        if "dentro_raio" in dfx.columns:
            dfx["dentro_raio"] = dfx["dentro_raio"].map(
                {True: "Dentro", False: "Fora"}).fillna("—")
        # Cabeçalhos amigáveis
        ren = {"horario_comp": "Horário", "distancia_km": "Erro (km)",
               "raio_km": "Raio Sistema (km)", "dentro_raio": "No Raio?",
               "estimada_comp": "Tipo Posição", "gps_valido_comp": "GPS Válido?",
               "lat_comp": "Latitude", "lon_comp": "Longitude",
               "_tech_comp": "Tecnologia", "_operadora_comp": "Operadora",
               "_banda_comp": "Banda", "endereco_comp": "Endereço Estimado"}
        dfx = dfx.rename(columns=ren)
        for c in dfx.columns:
            if "horario" in c:
                dfx[c] = dfx[c].astype(str)
        _escreve_df(ws3, dfx)
        n = len(dfx)
        if "Erro (km)" in dfx.columns:
            cdist = list(dfx.columns).index("Erro (km)") + 1
            line = LineChart(); line.title = "Erro de Posição ao Longo do Tempo (km)"
            line.style = 12
            data = Reference(ws3, min_col=cdist, min_row=1, max_row=n + 1)
            line.add_data(data, titles_from_data=True)
            line.height = 8; line.width = 20
            ws3.add_chart(line, f"{chr(65 + len(dfx.columns) + 1)}2")

    # ── Aba GLOSSÁRIO (para leitura sem a plataforma) ──
    _aba_glossario(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
