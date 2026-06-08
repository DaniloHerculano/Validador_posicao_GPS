"""
export.py — Geração de planilha Excel multi-aba COM gráficos nativos e CSV.
"""
import io
import re
import pandas as pd
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Cores Stoneridge
C_RED = "DD0933"
C_SLATE = "2C3946"
C_LIGHT = "D5E2F2"


def _nome_aba(nome: str) -> str:
    limpo = re.sub(r'[\[\]\:\*\?\/\\]', '_', nome)
    return (limpo[:28] or "comp")[:31]


def nome_arquivo_seguro(nome: str) -> str:
    return re.sub(r'[^A-Za-z0-9_.-]', '_', nome)


def _estiliza_header(ws, ncols, nrow=1):
    fill = PatternFill("solid", fgColor=C_SLATE)
    font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    thin = Side(style="thin", color="DCE4EE")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=nrow, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)


def _escreve_df(ws, df, start_row=1):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _estiliza_header(ws, len(df.columns), start_row)
    # largura automática simples
    for i, col in enumerate(df.columns, 1):
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].head(50)])
        ws.column_dimensions[ws.cell(row=1, column=i).coordinate[:1]].width = min(maxlen + 3, 45)


def gerar_excel(df_resumo, resultados: dict, dados: list,
                raios=(1.0, 3.0, 5.0)) -> bytes:
    """
    Excel com:
      - Resumo (tabela + gráfico de barras de erro médio + % por raio)
      - Rede_Bateria (consolidado + gráfico de pizza de tecnologia da referência)
      - Uma aba por equipamento sincronizado (tabela + linha de erro no tempo)
    """
    from openpyxl import Workbook
    wb = Workbook()
    raio1, raio2, raio3 = raios

    # ── Aba RESUMO ──
    ws = wb.active
    ws.title = "Resumo"
    if df_resumo is not None and len(df_resumo) > 0:
        _escreve_df(ws, df_resumo)
        n = len(df_resumo)

        # Gráfico de barras: erro médio (procura coluna que contenha "Erro Médio")
        col_erro = next((i for i, c in enumerate(df_resumo.columns, 1)
                         if "Erro Médio" in str(c)), None)
        col_equip = next((i for i, c in enumerate(df_resumo.columns, 1)
                          if "Equipamento" in str(c)), 1)
        if col_erro:
            bar = BarChart()
            bar.title = "Erro Médio por Equipamento (km)"
            bar.type = "col"
            bar.style = 10
            data = Reference(ws, min_col=col_erro, min_row=1, max_row=n + 1)
            cats = Reference(ws, min_col=col_equip, min_row=2, max_row=n + 1)
            bar.add_data(data, titles_from_data=True)
            bar.set_categories(cats)
            bar.height = 8; bar.width = 16
            ws.add_chart(bar, f"A{n + 4}")

        # Gráfico de barras agrupadas: % por raio
        cols_pct = [i for i, c in enumerate(df_resumo.columns, 1) if "%" in str(c)]
        if cols_pct:
            bar2 = BarChart()
            bar2.title = "% Dentro de Cada Raio"
            bar2.type = "col"; bar2.grouping = "clustered"; bar2.style = 12
            data = Reference(ws, min_col=min(cols_pct), max_col=max(cols_pct),
                             min_row=1, max_row=n + 1)
            cats = Reference(ws, min_col=col_equip, min_row=2, max_row=n + 1)
            bar2.add_data(data, titles_from_data=True)
            bar2.set_categories(cats)
            bar2.height = 8; bar2.width = 16
            ws.add_chart(bar2, f"A{n + 22}")

    # ── Aba REDE & BATERIA ──
    linhas = []
    for d in dados:
        df = d["df"]
        total = len(df)
        bat = df["_bateria_pct"].dropna() if "_bateria_pct" in df.columns else pd.Series(dtype=float)
        tem_tech = "_tech" in df.columns
        op = (df["_operadora"].mode().iloc[0]
              if "_operadora" in df.columns and df["_operadora"].notna().any() else "—")
        linhas.append({
            "Modelo": d.get("modelo", "—"), "PIN": d["pin"], "Tipo": d["tipo"],
            "Fonte": d.get("fonte", "—"), "Registros": total,
            "4G": int(df["_tech"].str.startswith("4G").sum()) if tem_tech else 0,
            "3G": int(df["_tech"].str.startswith("3G").sum()) if tem_tech else 0,
            "2G": int(df["_tech"].str.startswith("2G").sum()) if tem_tech else 0,
            "Operadora": op,
            "Bateria Início (%)": bat.iloc[0] if len(bat) else None,
            "Bateria Fim (%)": bat.iloc[-1] if len(bat) else None,
        })
    if linhas:
        ws2 = wb.create_sheet("Rede_Bateria")
        df_rb = pd.DataFrame(linhas)
        _escreve_df(ws2, df_rb)
        n = len(df_rb)
        # Barras empilhadas 2G/3G/4G por equipamento
        c4 = list(df_rb.columns).index("4G") + 1
        c2 = list(df_rb.columns).index("2G") + 1
        cmod = list(df_rb.columns).index("Modelo") + 1
        bar = BarChart(); bar.title = "Distribuição de Tecnologia por Equipamento"
        bar.type = "col"; bar.grouping = "stacked"; bar.overlap = 100; bar.style = 10
        data = Reference(ws2, min_col=min(c4, c2), max_col=max(c4, c2), min_row=1, max_row=n + 1)
        cats = Reference(ws2, min_col=cmod, min_row=2, max_row=n + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 9; bar.width = 18
        ws2.add_chart(bar, f"A{n + 4}")

    # ── Abas por equipamento sincronizado ──
    for nome, df in resultados.items():
        if len(df) == 0:
            continue
        ws3 = wb.create_sheet(f"Sync_{_nome_aba(nome)}"[:31])
        # Selecionar colunas legíveis
        cols = [c for c in ["horario_comp", "distancia_km", "lat_comp", "lon_comp",
                            "_tech_comp", "_operadora_comp", "endereco_comp"] if c in df.columns]
        dfx = df[cols].copy()
        for c in dfx.columns:
            if "horario" in c:
                dfx[c] = dfx[c].astype(str)
        _escreve_df(ws3, dfx)
        n = len(dfx)
        # Linha: erro ao longo do tempo
        if "distancia_km" in dfx.columns:
            cdist = list(dfx.columns).index("distancia_km") + 1
            line = LineChart(); line.title = "Erro de Posição ao Longo do Tempo (km)"
            line.style = 12
            data = Reference(ws3, min_col=cdist, min_row=1, max_row=n + 1)
            line.add_data(data, titles_from_data=True)
            line.height = 8; line.width = 20
            ws3.add_chart(line, f"{chr(65 + len(cols) + 1)}2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
